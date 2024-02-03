
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
	def __init__(self):
		super().__init__()
		self.layout_config()
		self.appearence()
		self.todo_systema()
		
		
	def layout_config(self):
		self.title("Cadastro de Atletas")
		self.geometry("900x600")
		
	def appearence(self):
		self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"]).place(x=50, y=510)
		self.opt_apm = ctk.CTkOptionMenu(self, values=["Dark", "Light"], command=self.change_apm).place(x=50, y=540)
		
	def todo_systema(self):
		frame = ctk.CTkFrame(self, width=900, height=70, corner_radius=0, bg_color="blue", fg_color="blue")
		frame.place(x=0, y=10)
		title = ctk.CTkLabel(frame, text="BANCO DE DADOS", font=("Century Gothic bold", 28), text_color="#fff")
		title.place(x=340, y=20)
		span = ctk.CTkLabel(self, text="", font=("Century Gothic bold", 17), text_color=["#fff", "#fff"]).place(x=370, y=80)
		
		ficheiro = pathlib.Path("Atletas.xlsx")
	
		if ficheiro.exists():
			pass
		else:
			ficheiro=Workbook()
			folha=ficheiro.active
			folha['A1']="Atleta"
			folha['B1']="Idade"
			folha['C1']="Clube"
			folha['D1']="Observaçoes"
			folha['E1']="Link"
			folha['F1']="Situação"
			
			ficheiro.save("Atletas.xlsx")
			
			
		def submit():
			
			
			
			#pegando os dados dos entrys
			name = name_value.get()
			age =  age_value.get()
			link = link_value.get()
			club = club_value.get()
			sit = sit_combobox.get()
			obs = obs_entry.get(0.0, END)
			
			#if (name =="" or age =="" or club =="" or link == ""):
			
				#messagebox.showerror("Sistema", "ERRO!\nIndique a situaçao do atleta!")
			ficheiro = openpyxl.load_workbook('Atletas.xlsx')
			folha = ficheiro.active
			folha.cell(column=1, row=folha.max_row+1, value=name)
			folha.cell(column=2, row=folha.max_row, value=age)
			folha.cell(column=3, row=folha.max_row, value=club)
			folha.cell(column=4, row=folha.max_row, value=obs)
			folha.cell(column=5, row=folha.max_row, value=link)
			folha.cell(column=6, row=folha.max_row, value=sit)
			
			ficheiro.save(r'Atletas.xlsx')
			messagebox.showinfo("Sistema", "Atletas cadastrado com sucesso!")
			
			
			
		def clear():
			name_value.set("")
			age_value.set("")
			link_value.set("")
			club_value.set("")
			obs_entry.delete(0.0, END)
			
			
		#texts variables
		name_value = StringVar()
		age_value = StringVar()
		link_value = StringVar()
		club_value = StringVar()
		
		#Entrys
		name_entry = ctk.CTkEntry(self, width=450, textvariable=name_value, font=("Century Gothic bold", 18), fg_color="transparent")
		age_entry = ctk.CTkEntry(self, width=220, textvariable=age_value, font=("Century Gothic bold", 18), fg_color="transparent")
		link_entry = ctk.CTkEntry(self, width=450, textvariable=link_value, font=("Century Gothic bold", 18), fg_color="transparent")
		club_entry = ctk.CTkEntry(self, width=220, textvariable=club_value, font=("Century Gothic bold", 18), fg_color="transparent")
		
		#Combobox
		sit_combobox = ctk.CTkComboBox(self, values=["Livre", "Emprestado", "Contrato"], font=("Century Gothic bold", 14), width=160)
		sit_combobox.set("Contrato")
		
		
		#Entrada de Observaçoes
		obs_entry = ctk.CTkTextbox(self, width=800, font=("arial", 19), border_color="#aaa", border_width=2, fg_color="transparent")
		
		
		#labels
		lb_name = ctk.CTkLabel(self, text="NOME", font=("Century Gothic bold", 17), text_color=["#000", "#fff"])
		lb_age = ctk.CTkLabel(self, text="IDADE", font=("Century Gothic bold", 17), text_color=["#000", "#fff"])
		lb_link = ctk.CTkLabel(self, text="TRASNFERMARKT", font=("Century Gothic bold", 17), text_color=["#000", "#fff"])
		lb_club = ctk.CTkLabel(self, text="CLUBE", font=("Century Gothic bold", 17), text_color=["#000", "#fff"])
		lb_obs = ctk.CTkLabel(self, text="ANALISES", font=("Century Gothic bold", 17), text_color=["#000", "#fff"])
		lb_sit = ctk.CTkLabel(self, text="SITUAÇÃO", font=("Century Gothic bold", 17), text_color=["#000", "#fff"])
		
		btn_submit = ctk.CTkButton(self, text="Salvar Dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=280, y=540)
		btn_submit = ctk.CTkButton(self, text="Apagar Dados".upper(), command=clear, fg_color="#151", hover_color="#333").place(x=480, y=540)
		
		#Posicionando na Janela
		lb_name.place(x=50, y=120)
		name_entry.place(x=50, y=150)
		
		lb_age.place(x=630, y=120)
		age_entry.place(x=630, y=150)
		
		lb_link.place(x=50, y=190)
		link_entry.place(x=50, y=220)
		
		lb_club.place(x=630, y=190)
		club_entry.place(x=630, y=220)
		
		lb_obs.place(x=50, y=260)
		obs_entry.place(x=50, y=290)
		
		lb_sit.place(x=690, y=510)
		sit_combobox.place(x=690, y=540)
		
		
	def change_apm(self, nova_aparencia):
		ctk.set_appearance_mode(nova_aparencia)
		
if __name__=="__main__":
	app = App()
	app.mainloop()


input()